from flask import Flask, request, jsonify, send_file, render_template, session
from flask_cors import CORS
import pandas as pd
import numpy as np
from scipy.stats import gaussian_kde
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from modeliai import RandomForestImputer
try:
    from modeliai import XGBoostImputer
    XGBOOST_IMPUTER_AVAILABLE = True
except ImportError:
    XGBOOST_IMPUTER_AVAILABLE = False
import io
import base64
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.graph_objs as go
import plotly.utils
import json
import os
from werkzeug.utils import secure_filename
import mysql.connector
from mysql.connector import Error
from datetime import datetime
import uuid
import json as json_module
from pathlib import Path
import smtplib
import psutil
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import formatdate, make_msgid
from email.mime.base import MIMEBase
from email import encoders
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import tempfile

# Load .env file for local development
def load_env():
    """Load environment variables from .env file"""
    env_path = Path('.') / '.env'
    if env_path.exists():
        print("Loading environment variables from .env file...")
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    os.environ[key] = value
                    print(f"Loaded: {key}")
    else:
        print("No .env file found, using system environment variables")

# Load environment variables at startup
load_env()

app = Flask(__name__)
CORS(app)

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-magistro-darbas-2025')
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # Disable cache for static files

# Disable template caching in development
@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# MySQL Configuration - only works in production with environment variables
# No default values for security - requires proper environment setup
DB_CONFIG = {
    'host': os.environ.get('MYSQL_HOST'),
    'user': os.environ.get('MYSQL_USER'),
    'password': os.environ.get('MYSQL_PASSWORD'),
    'database': os.environ.get('MYSQL_DATABASE'),
    'connect_timeout': 10,
    'ssl_disabled': False,
    'autocommit': True,
    'charset': 'utf8mb4',
    'collation': 'utf8mb4_unicode_ci',
    'client_flags': [mysql.connector.constants.ClientFlag.MULTI_STATEMENTS]
}

# Email Configuration
EMAIL_CONFIG = {
    'smtp_server': os.environ.get('SMTP_SERVER'),
    'smtp_port': int(os.environ.get('SMTP_PORT', 587)),
    'smtp_username': os.environ.get('SMTP_USERNAME'),
    'smtp_password': os.environ.get('SMTP_PASSWORD'),
    'from_email': os.environ.get('SMTP_FROM_EMAIL'),
    'from_name': os.environ.get('SMTP_FROM_NAME', 'Duomenų Analizės Sistema')
}

# Check if email is configured
EMAIL_ENABLED = all([
    EMAIL_CONFIG['smtp_server'],
    EMAIL_CONFIG['smtp_username'],
    EMAIL_CONFIG['smtp_password'],
    EMAIL_CONFIG['from_email']
])

# Check if all required MySQL environment variables are set
MYSQL_ENABLED = all([
    os.environ.get('MYSQL_HOST'),
    os.environ.get('MYSQL_USER'),
    os.environ.get('MYSQL_PASSWORD'),
    os.environ.get('MYSQL_DATABASE')
])

print(f"MySQL enabled: {MYSQL_ENABLED}")
if not MYSQL_ENABLED:
    print("MySQL environment variables not set - comments feature will be disabled")
    print("Required: MYSQL_HOST, MYSQL_USER, MYSQL_PASSWORD, MYSQL_DATABASE")

# Database connection manager
class DatabaseManager:
    def __init__(self, config, enabled=True):
        self.config = config
        self.connection = None
        self.enabled = enabled

    def get_connection(self):
        if not self.enabled:
            print("MySQL is disabled - environment variables not configured")
            return None

        try:
            # Always create a fresh connection for each request to avoid stale connection issues
            # This is simpler and more reliable than connection pooling for this use case
            if self.connection is not None:
                try:
                    self.connection.close()
                except:
                    pass

            self.connection = mysql.connector.connect(**self.config)
            return self.connection
        except Error as e:
            print(f"Database connection error: {e}")
            self.connection = None
            return None

    def close_connection(self):
        if self.connection and self.connection.is_connected():
            self.connection.close()

# Initialize database manager - only if MySQL is properly configured
db_manager = DatabaseManager(DB_CONFIG, enabled=MYSQL_ENABLED)

# Database initialization functions
def init_database_tables():
    """Initialize all required database tables"""
    if not MYSQL_ENABLED:
        return

    try:
        connection = db_manager.get_connection()
        if not connection:
            return

        cursor = connection.cursor()

        # Create imputacijos_rezultatai table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS imputacijos_rezultatai (
                id INT AUTO_INCREMENT PRIMARY KEY,
                rezultato_id VARCHAR(36) UNIQUE NOT NULL,
                originalus_failas VARCHAR(255) NOT NULL,
                imputuotas_failas VARCHAR(255) NOT NULL,
                modelio_tipas ENUM('random_forest', 'xgboost') NOT NULL,
                n_estimators INT NOT NULL,
                random_state INT NOT NULL,
                max_depth INT DEFAULT NULL,
                learning_rate FLOAT DEFAULT NULL,
                originalus_trukstamu_kiekis INT NOT NULL,
                imputuotas_trukstamu_kiekis INT NOT NULL,
                apdorotu_stulpeliu_kiekis INT NOT NULL,
                modelio_metrikos JSON,
                pozymiu_svarba JSON,
                grafikai JSON,
                test_predictions JSON,
                sukurimo_data DATETIME DEFAULT CURRENT_TIMESTAMP,
                busena ENUM('vykdomas', 'baigtas', 'klaida') DEFAULT 'vykdomas'
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        # Add test_predictions column if it doesn't exist (for existing tables)
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'imputacijos_rezultatai'
            AND COLUMN_NAME = 'test_predictions'
        """)
        result = cursor.fetchone()
        if result[0] == 0:  # Use index instead of key since cursor is not dictionary type
            cursor.execute("""
                ALTER TABLE imputacijos_rezultatai
                ADD COLUMN test_predictions JSON AFTER grafikai
            """)

        # Create rezultatu_komentarai table (separate from general komentarai)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS rezultatu_komentarai (
                id INT AUTO_INCREMENT PRIMARY KEY,
                komentaro_id VARCHAR(36) UNIQUE NOT NULL,
                rezultato_id VARCHAR(36) NOT NULL,
                vardas VARCHAR(100) NOT NULL,
                komentaras TEXT NOT NULL,
                sukurimo_data DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_rezultato_id (rezultato_id),
                FOREIGN KEY (rezultato_id) REFERENCES imputacijos_rezultatai(rezultato_id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        # Add max_depth and learning_rate columns if they don't exist (for existing tables)
        try:
            cursor.execute("""
                ALTER TABLE imputacijos_rezultatai
                ADD COLUMN max_depth INT DEFAULT NULL AFTER random_state
            """)
        except:
            pass  # Column already exists

        try:
            cursor.execute("""
                ALTER TABLE imputacijos_rezultatai
                ADD COLUMN learning_rate FLOAT DEFAULT NULL AFTER max_depth
            """)
        except:
            pass  # Column already exists

        connection.commit()
        cursor.close()
        db_manager.close_connection()
        print("Database tables initialized successfully")

    except Error as e:
        print(f"Error initializing database tables: {e}")

# Initialize tables on startup
init_database_tables()

# MLImputer wrapper klasė, kuri naudoja atskirų modelių failus
class MLImputer:
    """Wrapper klasė, kuri pasirenka tinkamą modelį pagal model_type"""

    def __init__(self, model_type='random_forest', n_estimators=100, random_state=42,
                 max_depth=None, learning_rate=None, exclude_columns=None):
        self.model_type = model_type
        self.n_estimators = n_estimators
        self.random_state = random_state
        self.max_depth = max_depth
        self.learning_rate = learning_rate
        self.exclude_columns = exclude_columns or []

        # Inicijuojame tinkamą modelį
        if model_type == 'xgboost' and XGBOOST_IMPUTER_AVAILABLE:
            try:
                # XGBoost specific parameters
                xgb_params = {
                    'n_estimators': n_estimators,
                    'random_state': random_state,
                    'exclude_columns': self.exclude_columns
                }
                if max_depth is not None:
                    xgb_params['max_depth'] = max_depth
                if learning_rate is not None:
                    xgb_params['learning_rate'] = learning_rate

                self.imputer = XGBoostImputer(**xgb_params)
            except Exception as e:
                # Bet kokios klaidos - naudojame Random Forest
                rf_params = {
                    'n_estimators': n_estimators,
                    'random_state': random_state,
                    'exclude_columns': self.exclude_columns
                }
                if max_depth is not None:
                    rf_params['max_depth'] = max_depth
                self.imputer = RandomForestImputer(**rf_params)
                self.model_type = 'random_forest'
        else:
            # Naudojame Random Forest (arba jei XGBoost neprieinamas)
            if model_type == 'xgboost' and not XGBOOST_IMPUTER_AVAILABLE:
                self.model_type = 'random_forest'

            rf_params = {
                'n_estimators': n_estimators,
                'random_state': random_state,
                'exclude_columns': self.exclude_columns
            }
            if max_depth is not None:
                rf_params['max_depth'] = max_depth

            self.imputer = RandomForestImputer(**rf_params)
    
    def fit_transform(self, df):
        """Užpildo trūkstamas reikšmes naudojant pasirinktą modelį"""
        return self.imputer.fit_and_impute(df)
    
    def get_feature_importance(self):
        """Grąžina požymių svarbos informaciją"""
        return self.imputer.get_feature_importance()
    
    def get_model_metrics(self):
        """Grąžina modelio metrikas"""
        return self.imputer.get_model_metrics()

    def get_test_predictions(self):
        """Grąžina test predictions duomenis"""
        return self.imputer.get_test_predictions()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/imputacija')
def imputacija():
    return render_template('imputacija.html')

@app.route('/ikelti-duomenys')
def ikelti_duomenys():
    return render_template('ikelti_duomenys.html')

@app.route('/api/file-stats/<filename>')
def get_file_stats(filename):
    """Get statistics for a specific file"""
    try:
        filepath = os.path.join(UPLOAD_FOLDER, secure_filename(filename))
        if not os.path.exists(filepath):
            return jsonify({"error": "Failas nerastas"}), 404

        df = pd.read_csv(filepath)

        # Calculate missing values per column
        missing_values = {}
        for col in df.columns:
            missing_values[col] = int(df[col].isnull().sum())

        stats = {
            "rows": len(df),
            "columns": len(df.columns),
            "missing_values": missing_values
        }

        return jsonify({
            "stats": stats,
            "filename": filename
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/uploaded-files')
def get_uploaded_files():
    """Get list of all uploaded CSV files with statistics"""
    try:
        files_info = []

        # Get all CSV files from upload folder
        if os.path.exists(UPLOAD_FOLDER):
            for filename in os.listdir(UPLOAD_FOLDER):
                if filename.endswith('.csv'):
                    filepath = os.path.join(UPLOAD_FOLDER, filename)

                    try:
                        # Read CSV to get statistics
                        df = pd.read_csv(filepath)

                        # Calculate statistics
                        total_cells = len(df) * len(df.columns)
                        missing_count = df.isnull().sum().sum()
                        missing_percentage = round((missing_count / total_cells * 100), 2) if total_cells > 0 else 0

                        # Get file info
                        file_stat = os.stat(filepath)
                        file_size = file_stat.st_size
                        upload_date = datetime.fromtimestamp(file_stat.st_mtime)

                        # Check if file has time_period column
                        has_time_data = 'year' in df.columns
                        year_range = None
                        if has_time_data:
                            years = df['year'].dropna().unique()
                            if len(years) > 0:
                                year_range = f"{int(min(years))} - {int(max(years))}"

                        files_info.append({
                            'filename': filename,
                            'rows': len(df),
                            'columns': len(df.columns),
                            'missing_count': int(missing_count),
                            'missing_percentage': missing_percentage,
                            'file_size': file_size,
                            'file_size_mb': round(file_size / (1024 * 1024), 2),
                            'upload_date': upload_date.strftime('%Y-%m-%d %H:%M:%S'),
                            'has_time_data': has_time_data,
                            'year_range': year_range,
                            'column_names': df.columns.tolist()[:10]  # First 10 columns
                        })
                    except Exception as e:
                        print(f"Error reading file {filename}: {str(e)}")
                        continue

        # Sort by upload date (newest first)
        files_info.sort(key=lambda x: x['upload_date'], reverse=True)

        return jsonify({
            'files': files_info,
            'total_files': len(files_info)
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/apie')
def apie():
    return render_template('apie.html')

@app.route('/komentarai')
def komentarai():
    return render_template('komentarai.html')

@app.route('/rezultatai')
def rezultatai():
    return render_template('rezultatai.html')

@app.route('/rezultatai/<result_id>')
def rezultatas_detali(result_id):
    return render_template('rezultatas_detali.html')

@app.route('/palyginimas')
def palyginimas():
    return render_template('palyginimas.html')

@app.route('/api/rezultatai', methods=['GET'])
def get_rezultatai():
    """Get all imputation results"""
    if not MYSQL_ENABLED:
        return jsonify({
            "rezultatai": [],
            "message": "Rezultatų funkcija neprieinama - duomenų bazė nesukonfigūruota"
        })

    try:
        connection = db_manager.get_connection()
        if not connection:
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT rezultato_id, originalus_failas, imputuotas_failas,
                   modelio_tipas, n_estimators, random_state,
                   originalus_trukstamu_kiekis, imputuotas_trukstamu_kiekis,
                   apdorotu_stulpeliu_kiekis, sukurimo_data, busena
            FROM imputacijos_rezultatai
            ORDER BY sukurimo_data DESC
        """)
        rezultatai = cursor.fetchall()
        cursor.close()

        return jsonify({"rezultatai": rezultatai})

    except Error as e:
        return jsonify({"error": f"Duomenų bazės klaida: {str(e)}"}), 500

@app.route('/api/rezultatai/<result_id>', methods=['GET'])
def get_rezultatas(result_id):
    """Get detailed imputation result"""
    if not MYSQL_ENABLED:
        return jsonify({"error": "Rezultatų funkcija neprieinama - duomenų bazė nesukonfigūruota"}), 503

    cursor = None
    try:
        connection = db_manager.get_connection()
        if not connection:
            print(f"Failed to get database connection for result {result_id}")
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor(dictionary=True)

        # Use longer timeout for large responses
        if connection:
            try:
                cursor.execute("SET SESSION MAX_EXECUTION_TIME=60000")  # 60 seconds
            except:
                pass  # Ignore if not supported

        cursor.execute("""
            SELECT * FROM imputacijos_rezultatai
            WHERE rezultato_id = %s
        """, (result_id,))

        rezultatas = cursor.fetchone()

        if not rezultatas:
            print(f"Result not found: {result_id}")
            return jsonify({"error": "Rezultatas nerastas"}), 404

        # Parse JSON fields with error handling
        try:
            if rezultatas['modelio_metrikos']:
                rezultatas['modelio_metrikos'] = json_module.loads(rezultatas['modelio_metrikos'])
        except Exception as e:
            print(f"Error parsing modelio_metrikos: {e}")
            rezultatas['modelio_metrikos'] = None

        try:
            if rezultatas['pozymiu_svarba']:
                rezultatas['pozymiu_svarba'] = json_module.loads(rezultatas['pozymiu_svarba'])
        except Exception as e:
            print(f"Error parsing pozymiu_svarba: {e}")
            rezultatas['pozymiu_svarba'] = None

        try:
            if rezultatas['grafikai']:
                rezultatas['grafikai'] = json_module.loads(rezultatas['grafikai'])
        except Exception as e:
            print(f"Error parsing grafikai: {e}")
            rezultatas['grafikai'] = None

        return jsonify({"rezultatas": rezultatas})

    except Error as e:
        print(f"Database error in get_rezultatas: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Duomenų bazės klaida: {str(e)}"}), 500

    except Exception as e:
        print(f"Unexpected error in get_rezultatas: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Serverio klaida: {str(e)}"}), 500

    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass
        db_manager.close_connection()

@app.route('/api/comments/<result_id>', methods=['GET'])
def get_comments(result_id):
    """Get all comments for a specific result"""
    if not MYSQL_ENABLED:
        return jsonify({"comments": []}), 200  # Return empty array if DB not configured

    cursor = None
    try:
        connection = db_manager.get_connection()
        if not connection:
            print("Failed to get database connection for comments")
            return jsonify({"comments": [], "count": 0}), 200  # Return empty instead of error

        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT komentaro_id, rezultato_id, vardas, komentaras, sukurimo_data
            FROM rezultatu_komentarai
            WHERE rezultato_id = %s
            ORDER BY sukurimo_data DESC
        """, (result_id,))

        comments = cursor.fetchall()

        # Format dates
        for comment in comments:
            if comment['sukurimo_data']:
                comment['sukurimo_data'] = comment['sukurimo_data'].strftime('%Y-%m-%d %H:%M:%S')

        return jsonify({
            "comments": comments,
            "count": len(comments)
        })

    except Exception as e:
        print(f"Error fetching comments: {str(e)}")
        import traceback
        traceback.print_exc()
        # Return empty array instead of error to prevent UI breaking
        return jsonify({"comments": [], "count": 0}), 200

    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass

def generate_result_pdf(rezultatas):
    """Generate a detailed PDF report with all result information"""
    temp_image_files = []  # Keep track of temporary image files
    try:
        print(f"Starting PDF generation for result: {rezultatas.get('rezultato_id')}")

        # Register Unicode-supporting fonts
        try:
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.pdfbase import pdfmetrics
            import os

            # Try Windows fonts (Arial, Segoe UI) or system fonts
            font_paths = [
                ('C:/Windows/Fonts/arial.ttf', 'C:/Windows/Fonts/arialbd.ttf', 'Arial'),
                ('C:/Windows/Fonts/segoeui.ttf', 'C:/Windows/Fonts/segoeuib.ttf', 'SegoeUI'),
            ]

            font_name = None
            for regular_path, bold_path, name in font_paths:
                if os.path.exists(regular_path) and os.path.exists(bold_path):
                    try:
                        pdfmetrics.registerFont(TTFont(name, regular_path))
                        pdfmetrics.registerFont(TTFont(f'{name}-Bold', bold_path))
                        font_name = name
                        font_name_bold = f'{name}-Bold'
                        print(f"Using {name} fonts for Unicode support")
                        break
                    except Exception as e:
                        print(f"Failed to load {name}: {e}")
                        continue

            if not font_name:
                raise Exception("No Unicode fonts found")

        except Exception as e:
            # Fallback to Helvetica if no Unicode fonts available
            font_name = 'Helvetica'
            font_name_bold = 'Helvetica-Bold'
            print(f"Warning: Using Helvetica fonts (no Unicode support): {e}")

        # Create temporary PDF file
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf', mode='wb')
        pdf_path = temp_pdf.name
        temp_pdf.close()
        print(f"Created temporary PDF file: {pdf_path}")

        # Create PDF document
        doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                              topMargin=0.5*inch, bottomMargin=0.5*inch,
                              leftMargin=0.5*inch, rightMargin=0.5*inch)

        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()
        print("PDF document structure initialized")

        # Custom styles with Unicode font
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=font_name_bold,
            fontSize=20,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=20,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontName=font_name_bold,
            fontSize=14,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=12,
            spaceBefore=12
        )

        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10
        )

        # Title
        elements.append(Paragraph("Imputacijos Rezultato Ataskaita", title_style))
        elements.append(Spacer(1, 0.3*inch))

        # Basic Information
        elements.append(Paragraph("Pagrindinė informacija", heading_style))

        # Create a style for wrapping long text
        wrap_style = ParagraphStyle(
            'WrapStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=10,
            leading=12
        )

        basic_info = [
            ['Parametras', 'Reikšmė'],
            ['Rezultato ID', Paragraph(rezultatas['rezultato_id'], wrap_style)],
            ['Originalus failas', Paragraph(rezultatas['originalus_failas'], wrap_style)],
            ['Imputuotas failas', Paragraph(rezultatas['imputuotas_failas'], wrap_style)],
            ['Modelio tipas', 'Random Forest' if rezultatas['modelio_tipas'] == 'random_forest' else 'XGBoost'],
            ['N Estimators', str(rezultatas['n_estimators'])],
            ['Random State', str(rezultatas['random_state'])],
            ['Trūkstamų pradžioje', f"{rezultatas['originalus_trukstamu_kiekis']:,}"],
            ['Trūkstamų po apdorojimo', f"{rezultatas['imputuotas_trukstamu_kiekis']:,}"],
            ['Apdoroti stulpeliai', str(rezultatas['apdorotu_stulpeliu_kiekis'])],
            ['Sukūrimo data', rezultatas['sukurimo_data'].strftime('%Y-%m-%d %H:%M:%S')]
        ]

        # Add max_depth if available
        if rezultatas.get('max_depth') is not None:
            basic_info.append(['Max Depth', str(rezultatas['max_depth'])])

        # Add learning_rate if available
        if rezultatas.get('learning_rate') is not None:
            basic_info.append(['Learning Rate', str(rezultatas['learning_rate'])])

        basic_table = Table(basic_info, colWidths=[2.5*inch, 4*inch])
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (0, -1), font_name_bold),
            ('FONTNAME', (0, 1), (1, -1), font_name),
        ]))
        elements.append(basic_table)
        elements.append(Spacer(1, 0.3*inch))

        # Model Metrics
        print("Processing model metrics...")
        if rezultatas.get('modelio_metrikos'):
            try:
                metrics = json_module.loads(rezultatas['modelio_metrikos']) if isinstance(rezultatas['modelio_metrikos'], str) else rezultatas['modelio_metrikos']
                print(f"Parsed metrics: {type(metrics)}")

                # Filter regression metrics
                regression_metrics = {col: m for col, m in metrics.items()
                                    if m.get('model_type') in ['regression', 'synthetic_test'] and 'rmse' in m}
                print(f"Found {len(regression_metrics)} regression metrics")

                if regression_metrics:
                    elements.append(Paragraph("Modelio metrikos", heading_style))
            except Exception as e:
                print(f"Error parsing model metrics: {e}")
                regression_metrics = {}
        else:
            print("No model metrics found")
            regression_metrics = {}

        if regression_metrics:

                metrics_data = [['Stulpelis', 'nRMSE', 'R²', 'SMAPE (%)', 'nMAE', 'Imties dydis']]

                for col, m in regression_metrics.items():
                    metrics_data.append([
                        col,
                        f"{m.get('nrmse', 0):.4f}",
                        f"{m.get('r2', 0):.4f}",
                        f"{m.get('smape', 0):.2f}",
                        f"{m.get('nmae', 0):.4f}",
                        str(m.get('sample_size', 0))
                    ])

                metrics_table = Table(metrics_data, colWidths=[1.5*inch, 0.9*inch, 0.8*inch, 1*inch, 0.9*inch, 1*inch])
                metrics_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                    ('FONTNAME', (0, 1), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]))
                elements.append(metrics_table)
                elements.append(Spacer(1, 0.3*inch))

        # Feature Importance
        print("Processing feature importance...")
        if rezultatas.get('pozymiu_svarba'):
            try:
                importance = json_module.loads(rezultatas['pozymiu_svarba']) if isinstance(rezultatas['pozymiu_svarba'], str) else rezultatas['pozymiu_svarba']
                print(f"Parsed importance: {type(importance)}")

                if importance:
                    elements.append(PageBreak())
                    elements.append(Paragraph("Požymių svarba", heading_style))

                    for col, features in importance.items():
                        if features and isinstance(features, dict):
                            elements.append(Paragraph(f"<b>Užpildant stulpelį: {col}</b>", normal_style))
                            elements.append(Spacer(1, 0.1*inch))

                            # Sort and get top 10 features
                            sorted_features = sorted(features.items(), key=lambda x: x[1], reverse=True)[:10]

                            feature_data = [['Požymis', 'Svarba']]
                            for feature, value in sorted_features:
                                feature_data.append([feature, f"{value:.4f}"])

                            feature_table = Table(feature_data, colWidths=[4*inch, 2*inch])
                            feature_table.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                                ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                                ('FONTNAME', (0, 1), (-1, -1), font_name),
                                ('FONTSIZE', (0, 0), (-1, 0), 10),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                            ]))
                            elements.append(feature_table)
                            elements.append(Spacer(1, 0.2*inch))
            except Exception as e:
                print(f"Error parsing feature importance: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("No feature importance found")

        # Graphics/Plots
        print("Processing graphics/plots...")
        if rezultatas.get('grafikai'):
            try:
                plots = json_module.loads(rezultatas['grafikai']) if isinstance(rezultatas['grafikai'], str) else rezultatas['grafikai']
                print(f"Parsed plots: {type(plots)}, keys: {plots.keys() if plots else 'None'}")

                if plots:
                    elements.append(PageBreak())
                    elements.append(Paragraph("Grafinė analizė", heading_style))
                    elements.append(Spacer(1, 0.2*inch))

                    plot_titles = {
                        'scatter_predictions': 'Faktinių ir prognozuotų reikšmių palyginimas',
                        'performance_table': 'Modelių efektyvumo lentelė',
                        'rmse_comparison': 'nRMSE palyginimas',
                        'r2_comparison': 'R² palyginimas',
                        'mape_comparison': 'SMAPE palyginimas',
                        'radar_comparison': 'Bendrasis modelių palyginimas'
                    }

                    # Add all plots
                    for plot_key in ['scatter_predictions', 'performance_table', 'rmse_comparison', 'r2_comparison', 'mape_comparison', 'radar_comparison']:
                        if plot_key in plots:
                            plot_data = plots[plot_key]
                            title = plot_titles.get(plot_key, plot_key)
                            print(f"Adding plot: {plot_key}")

                            elements.append(Paragraph(f"<b>{title}</b>", normal_style))
                            elements.append(Spacer(1, 0.1*inch))

                            # Convert base64 to image
                            try:
                                img_data = base64.b64decode(plot_data)

                                # Create temporary image file
                                temp_img = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                                temp_img.write(img_data)
                                temp_img.close()

                                # Keep track of temp file for cleanup later
                                temp_image_files.append(temp_img.name)

                                # Add image to PDF
                                img = Image(temp_img.name, width=6.5*inch, height=4.5*inch)
                                elements.append(img)
                                elements.append(Spacer(1, 0.2*inch))

                                print(f"Successfully added plot: {plot_key}")
                            except Exception as e:
                                print(f"Error adding plot {plot_key}: {e}")
                                import traceback
                                traceback.print_exc()
                                continue
            except Exception as e:
                print(f"Error parsing plots: {e}")
                import traceback
                traceback.print_exc()
        else:
            print("No plots found")

        # Build PDF
        print(f"Building PDF with {len(elements)} elements...")
        doc.build(elements)
        print(f"PDF successfully created at: {pdf_path}")

        # Clean up temporary image files after PDF is built
        for temp_img_path in temp_image_files:
            try:
                if os.path.exists(temp_img_path):
                    os.unlink(temp_img_path)
                    print(f"Cleaned up temp image: {temp_img_path}")
            except Exception as e:
                print(f"Error cleaning up temp image {temp_img_path}: {e}")

        return pdf_path

    except Exception as e:
        print(f"Error generating PDF: {e}")
        import traceback
        traceback.print_exc()

        # Clean up temporary image files on error
        for temp_img_path in temp_image_files:
            try:
                if os.path.exists(temp_img_path):
                    os.unlink(temp_img_path)
            except:
                pass

        return None

@app.route('/api/send-result-email', methods=['POST'])
def send_result_email():
    """Send result details to email with PDF attachment"""
    if not EMAIL_ENABLED:
        return jsonify({"error": "El. pašto siuntimas nesukonfigūruotas"}), 503

    pdf_path = None
    try:
        data = request.get_json()
        result_id = data.get('result_id')
        recipient_email = data.get('email', '').strip()

        # Validation
        if not result_id:
            return jsonify({"error": "Rezultato ID yra privalomas"}), 400
        if not recipient_email:
            return jsonify({"error": "El. pašto adresas yra privalomas"}), 400

        # Basic email validation
        if '@' not in recipient_email or '.' not in recipient_email:
            return jsonify({"error": "Neteisingas el. pašto adresas"}), 400

        # Get result from database
        if not MYSQL_ENABLED:
            return jsonify({"error": "Duomenų bazė nesukonfigūruota"}), 503

        connection = db_manager.get_connection()
        if not connection:
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT * FROM imputacijos_rezultatai
            WHERE rezultato_id = %s
        """, (result_id,))

        rezultatas = cursor.fetchone()
        cursor.close()

        if not rezultatas:
            return jsonify({"error": "Rezultatas nerastas"}), 404

        # Generate PDF report
        print(f"Generating PDF report for result {result_id}...")
        pdf_path = generate_result_pdf(rezultatas)

        if not pdf_path:
            return jsonify({"error": "Nepavyko sugeneruoti PDF ataskaitos"}), 500

        # Create email
        msg = MIMEMultipart('mixed')
        msg['Subject'] = f"Imputacijos rezultatas: {rezultatas['originalus_failas']}"
        msg['From'] = EMAIL_CONFIG['from_email']
        msg['To'] = recipient_email
        msg['Reply-To'] = EMAIL_CONFIG['from_email']
        msg['Date'] = formatdate(localtime=True)
        msg['Message-ID'] = make_msgid(domain='reapi.lt')

        # Create HTML body
        html_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 10px 10px 0 0; }}
                .content {{ background: #f8f9fa; padding: 20px; border-radius: 0 0 10px 10px; }}
                .stat {{ background: white; padding: 15px; margin: 10px 0; border-left: 4px solid #667eea; border-radius: 5px; }}
                .stat-label {{ font-weight: bold; color: #667eea; }}
                .stat-value {{ font-size: 1.2em; color: #2c3e50; }}
                .footer {{ text-align: center; margin-top: 20px; color: #6c757d; font-size: 0.9em; }}
                .attachment-note {{ background: #fff3cd; padding: 15px; margin: 20px 0; border-left: 4px solid #ffc107; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2>📊 Imputacijos Rezultatas</h2>
                    <p>Duomenų analizės sistema</p>
                </div>
                <div class="content">
                    <h3>Rezultato detalės</h3>

                    <div class="stat">
                        <div class="stat-label">Rezultato ID:</div>
                        <div class="stat-value">{rezultatas['rezultato_id']}</div>
                    </div>

                    <div class="stat">
                        <div class="stat-label">Originalus failas:</div>
                        <div class="stat-value">{rezultatas['originalus_failas']}</div>
                    </div>

                    <div class="stat">
                        <div class="stat-label">Imputuotas failas:</div>
                        <div class="stat-value">{rezultatas['imputuotas_failas']}</div>
                    </div>

                    <div class="stat">
                        <div class="stat-label">Modelio tipas:</div>
                        <div class="stat-value">{'Random Forest' if rezultatas['modelio_tipas'] == 'random_forest' else 'XGBoost'}</div>
                    </div>

                    <div class="stat">
                        <div class="stat-label">Parametrai:</div>
                        <div class="stat-value">
                            n_estimators: {rezultatas['n_estimators']}<br>
                            random_state: {rezultatas['random_state']}<br>
                            {'max_depth: ' + str(rezultatas['max_depth']) + '<br>' if rezultatas.get('max_depth') else ''}
                            {'learning_rate: ' + str(rezultatas['learning_rate']) + '<br>' if rezultatas.get('learning_rate') else ''}
                        </div>
                    </div>

                    <div class="stat">
                        <div class="stat-label">Originalus trūkstamų kiekis:</div>
                        <div class="stat-value">{rezultatas['originalus_trukstamu_kiekis']:,}</div>
                    </div>

                    <div class="stat">
                        <div class="stat-label">Imputuotas trūkstamų kiekis:</div>
                        <div class="stat-value">{rezultatas['imputuotas_trukstamu_kiekis']:,}</div>
                    </div>

                    <div class="stat">
                        <div class="stat-label">Apdorotų stulpelių kiekis:</div>
                        <div class="stat-value">{rezultatas['apdorotu_stulpeliu_kiekis']}</div>
                    </div>

                    <div class="stat">
                        <div class="stat-label">Sukurimo data:</div>
                        <div class="stat-value">{rezultatas['sukurimo_data'].strftime('%Y-%m-%d %H:%M:%S')}</div>
                    </div>

                    <div class="stat">
                        <div class="stat-label">Būsena:</div>
                        <div class="stat-value">{rezultatas['busena'].capitalize()}</div>
                    </div>

                    <div class="attachment-note">
                        <strong>📎 Pridėtas failas:</strong><br>
                        Detalesnė ataskaita su modelio metrikomis, požymių svarba ir grafikais pridėta kaip PDF failas.
                    </div>
                </div>
                <div class="footer">
                    <p>Šis el. laiškas buvo sugeneruotas automatiškai.</p>
                    <p>© {datetime.now().year} Duomenų Analizės Sistema</p>
                </div>
            </div>
        </body>
        </html>
        """

        # Attach HTML body
        html_part = MIMEText(html_body, 'html')
        msg.attach(html_part)

        # Attach PDF file
        with open(pdf_path, 'rb') as pdf_file:
            pdf_attachment = MIMEBase('application', 'pdf')
            pdf_attachment.set_payload(pdf_file.read())
            encoders.encode_base64(pdf_attachment)

            # Create filename for attachment
            filename = f"rezultatas_{rezultatas['rezultato_id'][:8]}.pdf"
            pdf_attachment.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(pdf_attachment)

        # Send email
        print(f"Sending email to {recipient_email}...")
        with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['smtp_username'], EMAIL_CONFIG['smtp_password'])
            server.send_message(msg)

        print(f"Email sent successfully to {recipient_email}")

        return jsonify({
            "success": True,
            "message": f"Rezultatas su PDF ataskaita sėkmingai išsiųstas į {recipient_email}"
        }), 200

    except Exception as e:
        print(f"Error sending email: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Klaida siunčiant el. laišką: {str(e)}"}), 500

    finally:
        # Clean up temporary PDF file
        if pdf_path and os.path.exists(pdf_path):
            try:
                os.unlink(pdf_path)
                print(f"Cleaned up temporary PDF: {pdf_path}")
            except Exception as e:
                print(f"Error cleaning up PDF: {e}")

@app.route('/api/comments', methods=['POST'])
def create_comment():
    """Create a new comment for a result"""
    if not MYSQL_ENABLED:
        return jsonify({"error": "Komentarų funkcija neprieinama - duomenų bazė nesukonfigūruota"}), 503

    try:
        data = request.get_json()
        result_id = data.get('result_id')
        vardas = data.get('vardas', '').strip()
        komentaras = data.get('komentaras', '').strip()

        # Validation
        if not result_id:
            return jsonify({"error": "Rezultato ID yra privalomas"}), 400
        if not vardas or len(vardas) < 2:
            return jsonify({"error": "Vardas turi būti bent 2 simbolių ilgio"}), 400
        if not komentaras or len(komentaras) < 5:
            return jsonify({"error": "Komentaras turi būti bent 5 simbolių ilgio"}), 400
        if len(vardas) > 100:
            return jsonify({"error": "Vardas negali būti ilgesnis nei 100 simbolių"}), 400
        if len(komentaras) > 1000:
            return jsonify({"error": "Komentaras negali būti ilgesnis nei 1000 simbolių"}), 400

        connection = db_manager.get_connection()
        if not connection:
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor()

        # Insert comment into rezultatu_komentarai table
        komentaro_id = str(uuid.uuid4())
        cursor.execute("""
            INSERT INTO rezultatu_komentarai (komentaro_id, rezultato_id, vardas, komentaras, sukurimo_data)
            VALUES (%s, %s, %s, %s, NOW())
        """, (komentaro_id, result_id, vardas, komentaras))

        connection.commit()
        cursor.close()

        return jsonify({
            "success": True,
            "komentaro_id": komentaro_id,
            "message": "Komentaras sėkmingai paskelbtas"
        }), 201

    except Exception as e:
        print(f"Error creating comment: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/palyginimas', methods=['POST'])
def compare_results():
    """Compare multiple imputation results"""
    if not MYSQL_ENABLED:
        return jsonify({"error": "Palyginimo funkcija neprieinama - duomenų bazė nesukonfigūruota"}), 503

    try:
        data = request.get_json()
        result_ids = data.get('result_ids', [])

        if len(result_ids) < 2:
            return jsonify({"error": "Pasirinkite bent 2 rezultatus palyginimui"}), 400

        connection = db_manager.get_connection()
        if not connection:
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor(dictionary=True)

        # Get selected results
        placeholders = ','.join(['%s'] * len(result_ids))
        cursor.execute(f"""
            SELECT * FROM imputacijos_rezultatai
            WHERE rezultato_id IN ({placeholders})
            ORDER BY sukurimo_data DESC
        """, result_ids)

        rezultatai = cursor.fetchall()
        cursor.close()

        if len(rezultatai) != len(result_ids):
            return jsonify({"error": "Kai kurie rezultatai nerasti"}), 404

        # Parse JSON fields for each result (skip heavy graphics data for comparison)
        for rezultatas in rezultatai:
            if rezultatas['modelio_metrikos']:
                rezultatas['modelio_metrikos'] = json_module.loads(rezultatas['modelio_metrikos'])
            if rezultatas['pozymiu_svarba']:
                rezultatas['pozymiu_svarba'] = json_module.loads(rezultatas['pozymiu_svarba'])
            # Skip graphics for comparison to reduce data size
            rezultatas['grafikai'] = None

        # Generate comparison data
        comparison_data = generate_comparison_data(rezultatai)

        return jsonify({
            "rezultatai": rezultatai,
            "palyginimo_duomenys": comparison_data
        })

    except Error as e:
        return jsonify({"error": f"Duomenų bazės klaida: {str(e)}"}), 500

def generate_comparison_data(results):
    """Generate comparison analysis data"""
    comparison = {
        "bendroji_statistika": {},
        "modeliu_palyginimas": {},
        "efektyvumo_rodikliai": {},
        "rekomendacijos": []
    }

    # Basic statistics comparison
    comparison["bendroji_statistika"] = {
        "rezultatu_skaicius": len(results),
        "modeliu_tipai": list(set([r['modelio_tipas'] for r in results])),
        "bendras_trukstamu_uzpildymas": sum([r['originalus_trukstamu_kiekis'] - r['imputuotas_trukstamu_kiekis'] for r in results])
    }

    # Model comparison
    model_stats = {}
    for result in results:
        model_type = result['modelio_tipas']
        if model_type not in model_stats:
            model_stats[model_type] = {
                "kiekis": 0,
                "vidutinis_uzpildymas": 0,
                "vidutinis_nrmse": 0,
                "vidutinis_r2": 0,
                "vidutinis_laikas": 0
            }

        model_stats[model_type]["kiekis"] += 1
        uzpildymas = result['originalus_trukstamu_kiekis'] - result['imputuotas_trukstamu_kiekis']
        model_stats[model_type]["vidutinis_uzpildymas"] += uzpildymas

        # Calculate average metrics from model_metrics
        if result['modelio_metrikos']:
            nrmse_values = []
            r2_values = []
            for col_metrics in result['modelio_metrikos'].values():
                if col_metrics.get('nrmse'):
                    nrmse_values.append(col_metrics['nrmse'])
                if col_metrics.get('r2'):
                    r2_values.append(col_metrics['r2'])

            if nrmse_values:
                model_stats[model_type]["vidutinis_nrmse"] += sum(nrmse_values) / len(nrmse_values)
            if r2_values:
                model_stats[model_type]["vidutinis_r2"] += sum(r2_values) / len(r2_values)

    # Calculate averages
    for model_type, stats in model_stats.items():
        if stats["kiekis"] > 0:
            stats["vidutinis_uzpildymas"] /= stats["kiekis"]
            stats["vidutinis_nrmse"] /= stats["kiekis"]
            stats["vidutinis_r2"] /= stats["kiekis"]

    comparison["modeliu_palyginimas"] = model_stats

    # Efficiency indicators
    comparison["efektyvumo_rodikliai"] = calculate_efficiency_indicators(results)

    # Generate recommendations
    comparison["rekomendacijos"] = generate_recommendations(model_stats, results)

    return comparison

def calculate_efficiency_indicators(results):
    """Calculate efficiency indicators for comparison"""
    indicators = {
        "geriausias_uzpildymas": None,
        "geriausias_nrmse": None,
        "geriausias_r2": None,
        "greičiausias": None
    }

    best_filling = max(results, key=lambda r: r['originalus_trukstamu_kiekis'] - r['imputuotas_trukstamu_kiekis'])
    indicators["geriausias_uzpildymas"] = {
        "rezultato_id": best_filling['rezultato_id'],
        "modelis": best_filling['modelio_tipas'],
        "uzpildyta": best_filling['originalus_trukstamu_kiekis'] - best_filling['imputuotas_trukstamu_kiekis']
    }

    return indicators

def generate_recommendations(model_stats, results):
    """Generate recommendations based on comparison"""
    recommendations = []

    if len(model_stats) >= 2:
        rf_stats = model_stats.get('random_forest', {})
        xgb_stats = model_stats.get('xgboost', {})

        if rf_stats and xgb_stats:
            if rf_stats.get('vidutinis_r2', 0) > xgb_stats.get('vidutinis_r2', 0):
                recommendations.append({
                    "tipas": "modelio_pasirinkimas",
                    "tekstas": "Random Forest modelis rodo geresnį R² rezultatą ir yra rekomenduojamas tikslumui.",
                    "prioritetas": "aukštas"
                })
            else:
                recommendations.append({
                    "tipas": "modelio_pasirinkimas",
                    "tekstas": "XGBoost modelis rodo geresnį R² rezultatą ir yra rekomenduojamas tikslumui.",
                    "prioritetas": "aukštas"
                })

    if len(results) > 3:
        recommendations.append({
            "tipas": "duomenu_kokybe",
            "tekstas": "Rekomenduojama patikrinti duomenų kokybę - atlikta daug bandymų.",
            "prioritetas": "vidutinis"
        })

    return recommendations

@app.route('/api/rezultatai/<result_id>/koreliacijos-analize', methods=['GET'])
def get_rezultato_koreliacijos_analize(result_id):
    """
    Analyze correlation between missing data percentage and model accuracy for a specific result.
    Returns data for scatter plot showing relationship between missing values and R²/MAPE for this result.
    """
    if not MYSQL_ENABLED:
        return jsonify({"error": "Analizės funkcija neprieinama - duomenų bazė nesukonfigūruota"}), 503

    cursor = None
    try:
        connection = db_manager.get_connection()
        if not connection:
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor(dictionary=True)

        # Get specific result with metrics
        cursor.execute("""
            SELECT rezultato_id, originalus_failas, modelio_tipas,
                   originalus_trukstamu_kiekis, imputuotas_trukstamu_kiekis,
                   apdorotu_stulpeliu_kiekis, modelio_metrikos, sukurimo_data
            FROM imputacijos_rezultatai
            WHERE rezultato_id = %s AND busena = 'baigtas' AND modelio_metrikos IS NOT NULL
        """, (result_id,))

        rezultatas = cursor.fetchone()

        if not rezultatas:
            return jsonify({
                "data_points": [],
                "message": "Rezultatas nerastas arba neturi metrikų."
            })

        # Process result to extract correlation data
        data_points = []

        try:
            # Parse model metrics
            metrikos = json_module.loads(rezultatas['modelio_metrikos']) if rezultatas['modelio_metrikos'] else {}

            if not metrikos:
                return jsonify({
                    "data_points": [],
                    "message": "Rezultatas neturi metrikų."
                })

            # Calculate total data cells (approximation based on processed columns)
            apdorotu_stulpeliu = rezultatas['apdorotu_stulpeliu_kiekis']

            if apdorotu_stulpeliu == 0:
                return jsonify({
                    "data_points": [],
                    "message": "Nėra apdorotų stulpelių."
                })

            # For each column that was imputed, create a data point
            for column_name, column_metrics in metrikos.items():
                r2 = column_metrics.get('r2')
                mape = column_metrics.get('mape')
                rmse = column_metrics.get('rmse')
                mae = column_metrics.get('mae')

                # Only add if we have at least R² or MAPE
                if r2 is None and mape is None:
                    continue

                # Get per-column missing percentage from metrics if available
                # (New format - added in latest version)
                if 'missing_percentage' in column_metrics:
                    missing_percentage = column_metrics['missing_percentage']
                    missing_count = column_metrics.get('missing_count', 0)
                    total_rows = column_metrics.get('total_rows', None)
                else:
                    # Fallback for old results - use overall missing percentage as approximation
                    originalus_missing = rezultatas['originalus_trukstamu_kiekis']
                    imputuotas_missing = rezultatas['imputuotas_trukstamu_kiekis']
                    uzpildyta = originalus_missing - imputuotas_missing

                    # Estimate percentage (this is approximate - based on total missing)
                    missing_percentage = (originalus_missing / (originalus_missing + uzpildyta * apdorotu_stulpeliu)) * 100 if originalus_missing > 0 else 0
                    missing_count = None
                    total_rows = None

                data_point = {
                    'rezultato_id': rezultatas['rezultato_id'],
                    'failas': rezultatas['originalus_failas'],
                    'stulpelis': column_name,
                    'modelis': rezultatas['modelio_tipas'],
                    'missing_percentage': round(missing_percentage, 2),
                    'missing_count': missing_count,
                    'total_rows': total_rows,
                    'r2': round(r2, 4) if r2 is not None else None,
                    'mape': round(mape, 2) if mape is not None else None,
                    'rmse': round(rmse, 4) if rmse is not None else None,
                    'mae': round(mae, 4) if mae is not None else None,
                    'data': rezultatas['sukurimo_data'].strftime('%Y-%m-%d') if rezultatas['sukurimo_data'] else None
                }

                data_points.append(data_point)

        except Exception as e:
            print(f"Error processing result {result_id}: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({
                "data_points": [],
                "message": f"Klaida apdorojant rezultatą: {str(e)}"
            })

        return jsonify({
            "data_points": data_points,
            "total_points": len(data_points),
            "modelio_tipas": rezultatas['modelio_tipas'],
            "message": f"Surinkti {len(data_points)} duomenų taškai šiam rezultatui"
        })

    except Error as e:
        print(f"Database error in correlation analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Duomenų bazės klaida: {str(e)}"}), 500

    except Exception as e:
        print(f"Unexpected error in correlation analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Serverio klaida: {str(e)}"}), 500

    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass

@app.route('/api/rezultatai/<result_id>/rodikliai', methods=['GET'])
def get_rezultato_rodikliai(result_id):
    """
    Get list of all indicators (columns) that were imputed for a specific result.
    Returns list of column names with their metrics.
    """
    if not MYSQL_ENABLED:
        return jsonify({"error": "Funkcija neprieinama - duomenų bazė nesukonfigūruota"}), 503

    cursor = None
    try:
        connection = db_manager.get_connection()
        if not connection:
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor(dictionary=True)

        # Get specific result with metrics
        cursor.execute("""
            SELECT rezultato_id, originalus_failas, imputuotas_failas, modelio_metrikos
            FROM imputacijos_rezultatai
            WHERE rezultato_id = %s AND busena = 'baigtas'
        """, (result_id,))

        rezultatas = cursor.fetchone()

        if not rezultatas:
            return jsonify({"error": "Rezultatas nerastas"}), 404

        # Parse model metrics to get column names
        metrikos = json_module.loads(rezultatas['modelio_metrikos']) if rezultatas['modelio_metrikos'] else {}

        if not metrikos:
            return jsonify({"rodikliai": [], "message": "Rezultatas neturi metrikų"})

        # Create list of indicators with their basic info
        rodikliai = []
        for column_name, column_metrics in metrikos.items():
            rodikliai.append({
                'pavadinimas': column_name,
                'r2': round(column_metrics.get('r2', 0), 4) if column_metrics.get('r2') else None,
                'missing_count': column_metrics.get('missing_count', 0),
                'missing_percentage': column_metrics.get('missing_percentage', 0)
            })

        return jsonify({
            "rodikliai": rodikliai,
            "total": len(rodikliai)
        })

    except Exception as e:
        print(f"Error getting indicators: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Klaida: {str(e)}"}), 500

    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass

@app.route('/api/rezultatai/<result_id>/kde/<indicator_name>', methods=['GET'])
def get_rezultato_kde(result_id, indicator_name):
    """
    Calculate KDE for original and imputed data for a specific indicator.
    Returns KDE curves and overlap metric.
    """
    if not MYSQL_ENABLED:
        return jsonify({"error": "Funkcija neprieinama - duomenų bazė nesukonfigūruota"}), 503

    cursor = None
    try:
        connection = db_manager.get_connection()
        if not connection:
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor(dictionary=True)

        # Get result with test_predictions and model type
        cursor.execute("""
            SELECT rezultato_id, test_predictions, modelio_tipas
            FROM imputacijos_rezultatai
            WHERE rezultato_id = %s AND busena = 'baigtas'
        """, (result_id,))

        rezultatas = cursor.fetchone()

        if not rezultatas:
            return jsonify({"error": "Rezultatas nerastas"}), 404

        # Parse test_predictions
        test_predictions = json_module.loads(rezultatas['test_predictions']) if rezultatas['test_predictions'] else None

        if not test_predictions:
            return jsonify({"error": "Test predictions nerasti - pakartokite imputaciją su nauja versija"}), 404

        # Check if indicator exists in test predictions
        if indicator_name not in test_predictions:
            return jsonify({"error": f"Rodiklis '{indicator_name}' nerastas test predictions"}), 404

        # Get test data for this indicator (20% test set)
        indicator_test = test_predictions[indicator_name]

        if not indicator_test or 'y_true' not in indicator_test or 'y_pred' not in indicator_test:
            return jsonify({"error": f"Rodiklis '{indicator_name}' neturi test duomenų"}), 400

        # Original data = y_true from test set (tikrosios reikšmės)
        original_data = np.array(indicator_test['y_true'])

        # Imputed data = y_pred from test set (imputuotos reikšmės)
        imputed_data = np.array(indicator_test['y_pred'])

        if len(original_data) == 0:
            return jsonify({"error": f"Rodiklis '{indicator_name}' neturi test duomenų"}), 400

        if len(imputed_data) == 0:
            return jsonify({"error": f"Rodiklis '{indicator_name}' neturi prognozių"}), 400

        # Calculate statistics on test set only
        original_mean = float(np.mean(original_data))
        original_std = float(np.std(original_data))
        imputed_mean = float(np.mean(imputed_data))
        imputed_std = float(np.std(imputed_data))

        # Count - this is the test set size (20% of data)
        test_count = len(original_data)
        imputed_count = test_count  # All test values were "imputed" (predicted)

        # Calculate KDE
        try:
            # Create KDE for original data
            kde_original = gaussian_kde(original_data)

            # Create KDE for imputed data
            kde_imputed = gaussian_kde(imputed_data)

            # Create x range for plotting (from min to max of both datasets)
            x_min = min(original_data.min(), imputed_data.min())
            x_max = max(original_data.max(), imputed_data.max())
            x_range = x_max - x_min

            # Extend range by 10% on each side for better visualization
            x_min -= x_range * 0.1
            x_max += x_range * 0.1

            # Create 200 points for smooth curve
            x_values = np.linspace(x_min, x_max, 200)

            # Calculate KDE values
            kde_original_values = kde_original(x_values)
            kde_imputed_values = kde_imputed(x_values)

            # Calculate KDE overlap (integral of minimum of two KDEs)
            # This gives us a measure of similarity between distributions
            min_kde = np.minimum(kde_original_values, kde_imputed_values)
            # Use np.trapezoid for newer versions, fallback to np.trapz for older
            try:
                overlap = float(np.trapezoid(min_kde, x_values))
            except AttributeError:
                overlap = float(np.trapz(min_kde, x_values))

            # Prepare response
            # Get model type abbreviation for legend
            model_type = rezultatas['modelio_tipas']
            model_abbr = 'rf' if model_type == 'random_forest' else 'xgb'

            return jsonify({
                "indicator": indicator_name,
                "model_type": model_type,
                "model_abbr": model_abbr,
                "original": {
                    "x": x_values.tolist(),
                    "y": kde_original_values.tolist(),
                    "mean": original_mean,
                    "std": original_std,
                    "count": int(test_count)
                },
                "imputed": {
                    "x": x_values.tolist(),
                    "y": kde_imputed_values.tolist(),
                    "mean": imputed_mean,
                    "std": imputed_std,
                    "count": int(test_count)
                },
                "overlap": round(overlap, 4),
                "imputed_values_count": int(imputed_count),
                "test_set_size": int(test_count),
                "statistics": {
                    "original_mean": round(original_mean, 4),
                    "imputed_mean": round(imputed_mean, 4),
                    "mean_difference": round(abs(imputed_mean - original_mean), 4),
                    "original_std": round(original_std, 4),
                    "imputed_std": round(imputed_std, 4),
                    "test_set_percentage": 30  # Typically 20% for test set
                }
            })

        except Exception as kde_error:
            print(f"KDE calculation error: {str(kde_error)}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"Klaida skaičiuojant KDE: {str(kde_error)}"}), 500

    except Exception as e:
        print(f"Error in KDE endpoint: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Klaida: {str(e)}"}), 500

    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass

@app.route('/api/komentarai', methods=['GET'])
def get_komentarai():
    # Check if MySQL is enabled
    if not MYSQL_ENABLED:
        return jsonify({
            "komentarai": [],
            "message": "Komentarų funkcija neprieinama - duomenų bazė nesukonfigūruota"
        })

    try:
        connection = db_manager.get_connection()
        if not connection:
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, vardas, el_pastas, komentaras, sukurimo_data
            FROM komentarai
            ORDER BY sukurimo_data DESC
        """)
        komentarai = cursor.fetchall()
        cursor.close()

        return jsonify({"komentarai": komentarai})
    except Error as e:
        return jsonify({"error": f"Duomenų bazės klaida: {str(e)}"}), 500

@app.route('/api/komentarai', methods=['POST'])
def add_komentaras():
    # Check if MySQL is enabled
    if not MYSQL_ENABLED:
        return jsonify({
            "error": "Komentarų funkcija neprieinama - duomenų bazė nesukonfigūruota. Susisiekite su administratoriumi."
        }), 503

    try:
        data = request.get_json()
        vardas = data.get('vardas', '').strip()
        el_pastas = data.get('el_pastas', '').strip()
        komentaras = data.get('komentaras', '').strip()

        if not vardas or not komentaras:
            return jsonify({"error": "Vardas ir komentaras yra privalomi"}), 400

        connection = db_manager.get_connection()
        if not connection:
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS komentarai (
                id INT AUTO_INCREMENT PRIMARY KEY,
                vardas VARCHAR(100) NOT NULL,
                el_pastas VARCHAR(255),
                komentaras TEXT NOT NULL,
                sukurimo_data DATETIME DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        # Fix existing table charset if it exists
        try:
            cursor.execute("""
                ALTER TABLE komentarai
                CONVERT TO CHARACTER SET utf8mb4
                COLLATE utf8mb4_unicode_ci
            """)
        except:
            pass

        cursor.execute("""
            INSERT INTO komentarai (vardas, el_pastas, komentaras)
            VALUES (%s, %s, %s)
        """, (vardas, el_pastas if el_pastas else None, komentaras))

        connection.commit()
        cursor.close()

        return jsonify({"message": "Komentaras sėkmingai pridėtas"})
    except Error as e:
        return jsonify({"error": f"Duomenų bazės klaida: {str(e)}"}), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({"error": "Failas neįkeltas"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "Failas nepasirinktas"}), 400
        
        if file and file.filename.lower().endswith('.csv'):
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            
            # Read and analyze the CSV
            df = pd.read_csv(filepath)
            
            # Basic statistics
            stats = {
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': df.columns.tolist(),
                'data_types': df.dtypes.astype(str).to_dict(),
                'missing_values': df.isnull().sum().to_dict(),
                'missing_percentage': (df.isnull().sum() / len(df) * 100).round(2).to_dict()
            }
            
            return jsonify({
                "message": "Failas sėkmingai įkeltas",
                "filename": filename,
                "stats": stats
            })
        
        return jsonify({"error": "Prašome įkelti CSV failą"}), 400
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/analyze/<filename>')
def analyze_data(filename):
    try:
        filepath = os.path.join(UPLOAD_FOLDER, secure_filename(filename))
        if not os.path.exists(filepath):
            return jsonify({"error": "Failas nerastas"}), 404
        
        df = pd.read_csv(filepath)
        
        # Generate visualizations
        plots = {}
        
        # Missing values heatmap
        plt.figure(figsize=(12, 8))
        sns.heatmap(df.isnull(), yticklabels=False, cbar=True, cmap='viridis')
        plt.title('Trūkstamų reikšmių šilumos žemėlapis')
        plt.tight_layout()
        
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=80, bbox_inches='tight')
        img_buffer.seek(0)
        plots['missing_heatmap'] = base64.b64encode(img_buffer.getvalue()).decode()
        plt.close()
        
        # Missing values bar chart
        missing_counts = df.isnull().sum()
        missing_counts = missing_counts[missing_counts > 0]
        
        if len(missing_counts) > 0:
            plt.figure(figsize=(10, 6))
            missing_counts.plot(kind='bar')
            plt.title('Trūkstamų reikšmių kiekis pagal stulpelį')
            plt.xlabel('Stulpeliai')
            plt.ylabel('Trūkstamų reikšmių skaičius')
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=80, bbox_inches='tight')
            img_buffer.seek(0)
            plots['missing_bar'] = base64.b64encode(img_buffer.getvalue()).decode()
            plt.close()
        
        # Correlation matrix for numeric columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 1:
            plt.figure(figsize=(10, 8))
            correlation_matrix = df[numeric_cols].corr()
            sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', center=0, fmt='.2f')
            plt.title('Koreliacijos matrica')
            plt.tight_layout()
            
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=80, bbox_inches='tight')
            img_buffer.seek(0)
            plots['correlation'] = base64.b64encode(img_buffer.getvalue()).decode()
            plt.close()
        
        # Summary statistics
        summary_stats = df.describe().to_dict()
        
        # Missing values percentage by column
        missing_by_column = []
        for col in df.columns:
            missing_count = df[col].isnull().sum()
            filled_count = len(df) - missing_count
            missing_percent = round((missing_count / len(df)) * 100, 2)

            # Include columns with missing values OR special columns (time_period, geo)
            if missing_count > 0 or col in ['year', 'geo']:
                missing_by_column.append({
                    "column": col,
                    "missing_count": int(missing_count),
                    "filled_count": int(filled_count),
                    "missing_percentage": missing_percent,
                    "data_type": str(df[col].dtype)
                })

        # Sort: columns by missing percentage (highest first), then special columns (time_period, geo) at the end
        def sort_key(item):
            if item['column'] == 'year':
                return (1, 0)  # Second to last
            elif item['column'] == 'geo':
                return (1, 1)  # Last
            else:
                return (0, -item['missing_percentage'])  # Regular columns by missing percentage descending

        missing_by_column.sort(key=sort_key)

        # Analyze missing values over time (if time_period column exists)
        missing_over_time = None
        if 'year' in df.columns:
            # Get all numeric/indicator columns (exclude geo and year)
            indicator_columns = [col for col in df.columns
                               if col not in ['geo', 'year']
                               and df[col].dtype in ['float64', 'int64', 'float32', 'int32']]

            # Get unique years
            years = sorted(df['year'].dropna().unique().tolist())

            missing_over_time = {
                'years': years,
                'indicators': {},
                'total_by_year': {}
            }

            for year in years:
                year_data = df[df['year'] == year]
                total_missing = 0

                for indicator in indicator_columns:
                    if indicator not in missing_over_time['indicators']:
                        missing_over_time['indicators'][indicator] = []

                    missing_count = year_data[indicator].isnull().sum()
                    total_count = len(year_data)
                    missing_percentage = (missing_count / total_count * 100) if total_count > 0 else 0

                    missing_over_time['indicators'][indicator].append({
                        'year': year,
                        'missing_count': int(missing_count),
                        'total_count': int(total_count),
                        'missing_percentage': round(missing_percentage, 2)
                    })

                    total_missing += missing_count

                # Calculate total missing percentage for this year
                total_cells = len(year_data) * len(indicator_columns)
                missing_over_time['total_by_year'][year] = {
                    'missing_count': int(total_missing),
                    'total_cells': int(total_cells),
                    'missing_percentage': round((total_missing / total_cells * 100) if total_cells > 0 else 0, 2)
                }

        return jsonify({
            "plots": plots,
            "summary_stats": summary_stats,
            "missing_summary": {
                "total_missing": int(df.isnull().sum().sum()),
                "columns_with_missing": len(df.columns[df.isnull().any()]),
                "complete_rows": len(df.dropna()),
                "missing_percentage": round(df.isnull().sum().sum() / (len(df) * len(df.columns)) * 100, 2)
            },
            "missing_by_column": missing_by_column,
            "columns": df.columns.tolist(),
            "missing_over_time": missing_over_time
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get_csv_data/<filename>')
def get_csv_data(filename):
    """Get CSV data as JSON for map visualization"""
    try:
        filepath = os.path.join(UPLOAD_FOLDER, secure_filename(filename))
        if not os.path.exists(filepath):
            return jsonify({"error": "Failas nerastas"}), 404

        df = pd.read_csv(filepath)

        # Convert DataFrame to list of dictionaries
        # Replace NaN values with None for JSON serialization
        data = df.replace({np.nan: None}).to_dict('records')

        return jsonify(data)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

def save_imputation_result(filename, imputed_filename, model_type, n_estimators, random_state,
                          original_missing, imputed_missing, plots, feature_importance, model_metrics,
                          max_depth=None, learning_rate=None, test_predictions=None):
    """Save imputation result to database"""
    if not MYSQL_ENABLED:
        return None

    try:
        connection = db_manager.get_connection()
        if not connection:
            return None

        result_id = str(uuid.uuid4())

        cursor = connection.cursor()
        cursor.execute("""
            INSERT INTO imputacijos_rezultatai (
                rezultato_id, originalus_failas, imputuotas_failas,
                modelio_tipas, n_estimators, random_state, max_depth, learning_rate,
                originalus_trukstamu_kiekis, imputuotas_trukstamu_kiekis,
                apdorotu_stulpeliu_kiekis, modelio_metrikos,
                pozymiu_svarba, grafikai, test_predictions, busena
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            result_id, filename, imputed_filename, model_type,
            n_estimators, random_state, max_depth, learning_rate,
            original_missing, imputed_missing,
            len(feature_importance) if feature_importance else 0,
            json_module.dumps(model_metrics) if model_metrics else None,
            json_module.dumps(feature_importance) if feature_importance else None,
            json_module.dumps(plots) if plots else None,
            json_module.dumps(test_predictions) if test_predictions else None,
            'baigtas'
        ))

        connection.commit()
        cursor.close()

        return result_id

    except Error as e:
        print(f"Error saving imputation result: {e}")
        return None

@app.route('/impute/<filename>', methods=['POST'])
def impute_missing_values(filename):
    try:
        print(f"\n{'='*80}")
        print(f"Starting imputation for file: {filename}")
        print(f"{'='*80}")

        filepath = os.path.join(UPLOAD_FOLDER, secure_filename(filename))
        if not os.path.exists(filepath):
            return jsonify({"error": "Failas nerastas"}), 404

        print("Loading CSV file...")
        df = pd.read_csv(filepath)
        print(f"File loaded: {len(df)} rows, {len(df.columns)} columns")

        original_missing = df.isnull().sum().sum()
        print(f"Missing values: {original_missing}")

        if original_missing == 0:
            return jsonify({"message": "Nėra trūkstamų reikšmių užpildymui"}), 400

        # Get parameters from request
        data = request.get_json() or {}
        model_type = data.get('model_type', 'random_forest')
        n_estimators = data.get('n_estimators', 100)
        random_state = data.get('random_state', 42)
        max_depth = data.get('max_depth', None)  # None = automatic
        learning_rate = data.get('learning_rate', None)  # For XGBoost only
        exclude_columns = data.get('exclude_columns', [])  # Columns to exclude from imputation

        print(f"Model type: {model_type}, n_estimators: {n_estimators}")
        if exclude_columns:
            print(f"Excluded columns ({len(exclude_columns)}): {exclude_columns}")

        # Perform imputation with all parameters
        print("Starting imputation process...")
        imputer = MLImputer(
            model_type=model_type,
            n_estimators=n_estimators,
            random_state=random_state,
            max_depth=max_depth,
            learning_rate=learning_rate,
            exclude_columns=exclude_columns
        )
        df_imputed = imputer.fit_transform(df)
        print("Imputation completed!")

        # Generate unique filename for imputed data
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_id = str(uuid.uuid4())[:8]  # Short unique ID

        # Remove extension from original filename
        base_name = os.path.splitext(filename)[0]
        extension = os.path.splitext(filename)[1]

        # Create unique imputed filename
        imputed_filename = f"{base_name}_imputed_{model_type}_{timestamp}_{unique_id}{extension}"
        imputed_filepath = os.path.join(UPLOAD_FOLDER, imputed_filename)
        df_imputed.to_csv(imputed_filepath, index=False)

        # Generate comparison plots
        plots = {}

        # Before/After missing values comparison
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))

        # Original missing values
        missing_original = df.isnull().sum()
        missing_original = missing_original[missing_original > 0]
        if len(missing_original) > 0:
            missing_original.plot(kind='bar', ax=ax1)
            ax1.set_title('Trūkstamos reikšmės - originalūs duomenys')
            ax1.set_xlabel('Stulpeliai')
            ax1.set_ylabel('Trūkstamų kiekis')
            ax1.tick_params(axis='x', rotation=45)

        # After imputation (should be zero)
        missing_imputed = df_imputed.isnull().sum()
        missing_imputed = missing_imputed[missing_imputed > 0]
        if len(missing_imputed) > 0:
            missing_imputed.plot(kind='bar', ax=ax2)
        else:
            ax2.bar([], [])
        ax2.set_title('Trūkstamos reikšmės - po užpildymo')
        ax2.set_xlabel('Stulpeliai')
        ax2.set_ylabel('Trūkstamų kiekis')
        ax2.tick_params(axis='x', rotation=45)

        plt.tight_layout()
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=80, bbox_inches='tight')
        img_buffer.seek(0)
        plots['comparison'] = base64.b64encode(img_buffer.getvalue()).decode()
        plt.close()
        print("Comparison plot generated")

        # Feature importance plots - generate individual plot for each column
        print("Generating feature importance plots...")
        feature_importance = imputer.get_feature_importance()
        if feature_importance:
            print(f"Found {len(feature_importance)} columns with feature importance")
            plot_count = 0
            for col, importance in feature_importance.items():
                if importance:
                    plot_count += 1
                    print(f"  Generating plot {plot_count}/{len(feature_importance)} for {col}...")

                    # Sort by importance (descending)
                    sorted_importance = dict(sorted(importance.items(), key=lambda x: x[1], reverse=True))

                    # Take top 15 features for better visualization
                    top_features = dict(list(sorted_importance.items())[:15])
                    features = list(top_features.keys())
                    importances = list(top_features.values())

                    # Create individual plot for this column
                    fig, ax = plt.subplots(figsize=(10, max(6, len(features) * 0.4)))

                    # Horizontal bar chart
                    bars = ax.barh(range(len(features)), importances, color='#667eea')
                    ax.set_yticks(range(len(features)))
                    ax.set_yticklabels(features)
                    ax.set_xlabel('Svarba (Importance)', fontsize=11, fontweight='bold')
                    ax.set_title(f'Požymių svarba užpildant: {col}', fontsize=13, fontweight='bold', pad=15)

                    # Add value labels on bars
                    for i, (bar, val) in enumerate(zip(bars, importances)):
                        ax.text(val, i, f' {val:.4f}', va='center', fontsize=9)

                    ax.grid(axis='x', alpha=0.3, linestyle='--')
                    plt.tight_layout()

                    img_buffer = io.BytesIO()
                    plt.savefig(img_buffer, format='png', dpi=80, bbox_inches='tight')
                    img_buffer.seek(0)

                    # Store with column-specific key
                    plots[f'feature_importance_{col}'] = base64.b64encode(img_buffer.getvalue()).decode()
                    plt.close()
            print(f"Feature importance plots completed ({plot_count} plots)")
        else:
            print("No feature importance data found")

        # Generate model performance plots
        print("Generating model performance plots...")
        model_metrics = imputer.get_model_metrics()

        # Add per-column missing data information to metrics
        if model_metrics:
            missing_original = df.isnull().sum()
            total_rows = len(df)

            for column_name in model_metrics.keys():
                if column_name in df.columns:
                    missing_count = int(missing_original[column_name])
                    missing_percentage = round((missing_count / total_rows) * 100, 2)

                    # Add missing data info to each column's metrics
                    model_metrics[column_name]['missing_count'] = missing_count
                    model_metrics[column_name]['missing_percentage'] = missing_percentage
                    model_metrics[column_name]['total_rows'] = total_rows

        performance_plots = generate_model_performance_plots(model_metrics)
        plots.update(performance_plots)
        print("Model performance plots generated")

        # Generate scatter plots for actual vs predicted values
        print("Generating scatter plots...")
        test_predictions = imputer.get_test_predictions()
        scatter_plots = generate_scatter_plots(test_predictions, model_type)
        plots.update(scatter_plots)
        print("Scatter plots generated")

        # Convert test_predictions numpy arrays to lists for JSON serialization
        test_predictions_serializable = {}
        if test_predictions:
            for column_name, pred_data in test_predictions.items():
                # Check if pred_data is a dictionary with y_true and y_pred
                if isinstance(pred_data, dict) and 'y_true' in pred_data and 'y_pred' in pred_data:
                    test_predictions_serializable[column_name] = {
                        'y_true': pred_data['y_true'].tolist() if isinstance(pred_data['y_true'], np.ndarray) else pred_data['y_true'],
                        'y_pred': pred_data['y_pred'].tolist() if isinstance(pred_data['y_pred'], np.ndarray) else pred_data['y_pred'],
                        'test_indices': pred_data.get('test_indices', [])  # Include test indices if available
                    }
                else:
                    print(f"Warning: Skipping {column_name} - invalid test prediction format: {type(pred_data)}")
                    print(f"pred_data content: {pred_data}")

        # Save result to database (skip if disabled for testing)
        result_id = None
        SKIP_DB_SAVE = os.environ.get('SKIP_DB_SAVE', 'false').lower() == 'true'

        if not SKIP_DB_SAVE:
            print("Saving results to database...")
            try:
                result_id = save_imputation_result(
                    filename, imputed_filename, model_type, n_estimators, random_state,
                    int(original_missing), int(df_imputed.isnull().sum().sum()),
                    plots, feature_importance, model_metrics,
                    max_depth, learning_rate, test_predictions_serializable
                )
                if result_id:
                    print(f"Results saved with ID: {result_id}")
                else:
                    print("Warning: Failed to save results to database")
            except Exception as db_error:
                print(f"Database save failed: {db_error}")
                print("Continuing without database save...")
        else:
            print("Database save skipped (SKIP_DB_SAVE=true)")

        print("Preparing response data...")
        # Don't send plots in response - too large for big datasets
        # Plots are saved in database and will be loaded separately
        response_data = {
            "message": "Užpildymas sėkmingai baigtas",
            "imputed_filename": imputed_filename,
            "original_missing": int(original_missing),
            "imputed_missing": int(df_imputed.isnull().sum().sum()),
            "plots": {},  # Empty - load from DB instead
            "feature_importance": feature_importance,
            "model_metrics": model_metrics,
            "model_type": model_type
        }

        if result_id:
            response_data["result_id"] = result_id
            print(f"Results saved with ID: {result_id}")
            print("Plots stored in database - will be loaded separately")

        # Log response size
        import sys
        response_size = sys.getsizeof(str(response_data))
        print(f"Response data size (without plots): {response_size / 1024:.2f} KB")
        print(f"Number of plots generated: {len(plots)}")
        print(f"{'='*80}")
        print("Imputation completed successfully!")
        print(f"{'='*80}\n")

        return jsonify(response_data)

    except Exception as e:
        import traceback
        print(f"\n{'='*80}")
        print("ERROR during imputation:")
        print(f"{'='*80}")
        print(f"Error message: {str(e)}")
        print("Traceback:")
        traceback.print_exc()
        print(f"{'='*80}\n")
        return jsonify({"error": str(e)}), 500

@app.route('/download/<path:filename>')
def download_file(filename):
    try:
        # Use secure_filename but preserve the structure
        secure_name = secure_filename(filename)
        filepath = os.path.join(UPLOAD_FOLDER, secure_name)

        # Debug: print what we're looking for
        print(f"Looking for file: {filepath}")
        print(f"Original filename: {filename}")
        print(f"Secure filename: {secure_name}")

        if not os.path.exists(filepath):
            # Try without secure_filename in case it's mangling the name
            direct_path = os.path.join(UPLOAD_FOLDER, filename)
            print(f"Trying direct path: {direct_path}")

            if os.path.exists(direct_path):
                filepath = direct_path
            else:
                # List all files in upload folder for debugging
                import glob
                all_files = glob.glob(os.path.join(UPLOAD_FOLDER, "*"))
                print(f"Available files: {[os.path.basename(f) for f in all_files]}")
                return jsonify({"error": f"Failas nerastas: {filename}"}), 404

        return send_file(filepath, as_attachment=True)

    except Exception as e:
        print(f"Download error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/rezultatai/<result_id>/download-test-comparison', methods=['GET'])
def download_test_comparison_excel(result_id):
    """
    Generate and download Excel file with test set comparison.
    Works for both Random Forest and XGBoost models.
    Shows actual values (green) vs predicted values (yellow) for 20% test data.
    Format in cells: 118 (green) / 112.554654 (yellow background)
    All other imputed values are marked with yellow background.
    """
    if not MYSQL_ENABLED:
        return jsonify({"error": "Funkcija neprieinama - duomenų bazė nesukonfigūruota"}), 503

    cursor = None
    try:
        connection = db_manager.get_connection()
        if not connection:
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor(dictionary=True)

        # Get result data from database
        cursor.execute("""
            SELECT originalus_failas, imputuotas_failas, test_predictions, modelio_tipas
            FROM imputacijos_rezultatai
            WHERE rezultato_id = %s
        """, (result_id,))

        result = cursor.fetchone()

        if not result:
            return jsonify({"error": "Rezultatas nerastas"}), 404

        # Parse test_predictions JSON
        test_predictions = json_module.loads(result['test_predictions']) if result['test_predictions'] else {}

        if not test_predictions:
            return jsonify({"error": "Test predictions duomenys nerasti"}), 404

        # Load original and imputed dataframes
        original_filepath = os.path.join(UPLOAD_FOLDER, result['originalus_failas'])
        imputed_filepath = os.path.join(UPLOAD_FOLDER, result['imputuotas_failas'])

        if not os.path.exists(original_filepath):
            return jsonify({"error": "Originalus failas nerastas"}), 404
        if not os.path.exists(imputed_filepath):
            return jsonify({"error": "Imputuotas failas nerastas"}), 404

        df_original = pd.read_csv(original_filepath)
        df_imputed = pd.read_csv(imputed_filepath)

        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define styles
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        yellow_font = Font(color="9C6500", bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Define border style - thin gray borders
        thin_border = Border(
            left=Side(style='thin', color='A6A6A6'),
            right=Side(style='thin', color='A6A6A6'),
            top=Side(style='thin', color='A6A6A6'),
            bottom=Side(style='thin', color='A6A6A6')
        )

        # Identify which cells were originally missing (need to be imputed)
        missing_mask = df_original.isnull()

        # Create sheet with full data
        ws = wb.create_sheet("Test Set Comparison")

        # Write headers
        headers = list(df_imputed.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Write data
        for row_idx in range(len(df_imputed)):
            for col_idx, col_name in enumerate(headers, start=1):
                value = df_imputed.iloc[row_idx, col_idx - 1]
                cell = ws.cell(row=row_idx + 2, column=col_idx)

                # Check if this cell was originally missing
                was_missing = missing_mask.iloc[row_idx, col_idx - 1]

                # Check if this column has test predictions and this row is in test set
                if col_name in test_predictions:
                    pred_data = test_predictions[col_name]
                    test_indices = pred_data.get('test_indices', [])

                    if row_idx in test_indices:
                        # This is a TEST SET cell - show actual (green) / predicted (yellow)
                        test_pos = test_indices.index(row_idx)
                        y_true = pred_data['y_true'][test_pos]
                        y_pred = pred_data['y_pred'][test_pos]

                        # Format: actual (green) / predicted (yellow) using Rich Text
                        # Create rich text with different formatting for each part
                        true_text = f"{y_true:.6g}"
                        pred_text = f"{y_pred:.6f}"

                        # Green text for true value
                        green_inline_font = InlineFont(color="006100", b=True)
                        # Yellow/brown text for predicted value
                        yellow_inline_font = InlineFont(color="9C6500", b=True)

                        # Create rich text: green true value + separator + yellow predicted value
                        cell.value = CellRichText(
                            TextBlock(green_inline_font, true_text),
                            TextBlock(InlineFont(), " / "),
                            TextBlock(yellow_inline_font, pred_text)
                        )
                        # Light background for the whole cell
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.alignment = center_alignment
                        cell.border = thin_border
                    elif was_missing:
                        # This cell was imputed but NOT in test set - show only imputed value in yellow
                        cell.value = value
                        cell.fill = yellow_fill
                        cell.font = yellow_font
                        cell.border = thin_border
                    else:
                        # Original value (not missing)
                        cell.value = value
                        cell.border = thin_border
                elif was_missing:
                    # This cell was imputed - mark it yellow
                    cell.value = value
                    cell.fill = yellow_fill
                    cell.font = yellow_font
                    cell.border = thin_border
                else:
                    # Original value (not imputed)
                    cell.value = value
                    cell.border = thin_border

        # Adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = 20

        # Create legend sheet
        legend_ws = wb.create_sheet("Legenda", 0)
        legend_ws['A1'] = "Spalvų paaiškinimas"
        legend_ws['A1'].font = Font(bold=True, size=16)

        legend_ws['A3'] = "Pavyzdys"
        legend_ws['A3'].font = Font(bold=True, size=12)
        legend_ws['B3'] = "Ką reiškia"
        legend_ws['B3'].font = Font(bold=True, size=12)

        # Example with rich text
        legend_example_green = InlineFont(color="006100", b=True)
        legend_example_yellow = InlineFont(color="9C6500", b=True)
        legend_ws['A4'].value = CellRichText(
            TextBlock(legend_example_green, "118.0"),
            TextBlock(InlineFont(), " / "),
            TextBlock(legend_example_yellow, "112.554654")
        )
        legend_ws['A4'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        legend_ws['A4'].border = thin_border
        legend_ws['A4'].alignment = center_alignment
        legend_ws['B4'] = "Tikroji reikšmė / Modelio prognozė"

        legend_ws['A5'] = "112.554654"
        legend_ws['A5'].fill = yellow_fill
        legend_ws['A5'].font = yellow_font
        legend_ws['A5'].border = thin_border
        legend_ws['A5'].alignment = center_alignment
        legend_ws['B5'] = "Užpildyta reikšmė"

        legend_ws['A6'] = "95.8"
        legend_ws['A6'].border = thin_border
        legend_ws['A6'].alignment = center_alignment
        legend_ws['B6'] = "Originali reikšmė (nebuvo trūkstama)"

        legend_ws['A8'] = "Pastaba:"
        legend_ws['A8'].font = Font(bold=True, size=11)
        legend_ws['A9'] = "20% duomenų su žinomomis reikšmėmis naudojami modelio tikslumo vertinimui."
        legend_ws['A10'] = "Šiuose langeliuose matote tikrąją reikšmę ir modelio prognozę."

        legend_ws.column_dimensions['A'].width = 25
        legend_ws.column_dimensions['B'].width = 45

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"imputed_palyginimas_{result['modelio_tipas']}_{timestamp}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error generating Excel file: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Klaida generuojant Excel failą: {str(e)}"}), 500

    finally:
        if cursor:
            cursor.close()

@app.route('/api/rezultatai/<result_id>/download-original-xlsx', methods=['GET'])
def download_original_xlsx(result_id):
    """
    Download original CSV file converted to Excel format (.xlsx)
    """
    if not MYSQL_ENABLED:
        return jsonify({"error": "Funkcija neprieinama - duomenų bazė nesukonfigūruota"}), 503

    cursor = None
    try:
        connection = db_manager.get_connection()
        if not connection:
            return jsonify({"error": "Nepavyko prisijungti prie duomenų bazės"}), 500

        cursor = connection.cursor(dictionary=True)

        # Get result data from database
        cursor.execute("""
            SELECT originalus_failas
            FROM imputacijos_rezultatai
            WHERE rezultato_id = %s
        """, (result_id,))

        result = cursor.fetchone()

        if not result:
            return jsonify({"error": "Rezultatas nerastas"}), 404

        # Load original CSV file
        original_filepath = os.path.join(UPLOAD_FOLDER, result['originalus_failas'])

        if not os.path.exists(original_filepath):
            return jsonify({"error": "Originalus failas nerastas"}), 404

        df_original = pd.read_csv(original_filepath)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Original Data"

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='A6A6A6'),
            right=Side(style='thin', color='A6A6A6'),
            top=Side(style='thin', color='A6A6A6'),
            bottom=Side(style='thin', color='A6A6A6')
        )

        # Write headers
        headers = list(df_original.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Write data
        for row_idx in range(len(df_original)):
            for col_idx, col_name in enumerate(headers, start=1):
                value = df_original.iloc[row_idx, col_idx - 1]
                cell = ws.cell(row=row_idx + 2, column=col_idx)
                cell.value = value
                cell.border = thin_border

        # Adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = 15

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(result['originalus_failas'])[0]
        filename = f"{base_name}_original_{timestamp}.xlsx"

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error generating original Excel file: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Klaida generuojant Excel failą: {str(e)}"}), 500

    finally:
        if cursor:
            cursor.close()

def generate_scatter_plots(test_predictions, model_type):
    """Generate scatter plots for actual vs predicted values"""
    plots = {}

    if not test_predictions or len(test_predictions) == 0:
        return plots

    # Determine plot color based on model type
    plot_color = '#3498db' if model_type == 'random_forest' else '#e74c3c'
    model_name = 'Random Forest' if model_type == 'random_forest' else 'XGBoost'

    # Create grid for multiple scatter plots
    n_indicators = len(test_predictions)
    if n_indicators == 0:
        return plots

    # Calculate grid size
    n_cols = min(3, n_indicators)  # Max 3 columns
    n_rows = (n_indicators + n_cols - 1) // n_cols

    fig, axes = plt.subplots(n_rows, n_cols, figsize=(n_cols * 6, n_rows * 5))
    if n_indicators == 1:
        axes = [axes]
    elif n_rows == 1:
        axes = list(axes) if n_cols > 1 else [axes]
    else:
        axes = axes.flatten()

    for idx, (indicator, data) in enumerate(test_predictions.items()):
        ax = axes[idx]

        y_true = data['y_true']
        y_pred = data['y_pred']
        r2 = data['r2']
        mae = data.get('mae', np.mean(np.abs(y_true - y_pred)))
        rmse = data.get('rmse', np.sqrt(np.mean((y_true - y_pred)**2)))

        # Scatter plot
        ax.scatter(y_true, y_pred, alpha=0.6, edgecolors='k', s=60, color=plot_color)

        # Ideal line (y = x)
        min_val = min(y_true.min(), y_pred.min())
        max_val = max(y_true.max(), y_pred.max())
        ax.plot([min_val, max_val], [min_val, max_val], 'r--', lw=2, label='Ideali linija (y=x)')

        # Formatting
        ax.set_xlabel('Faktinės reikšmės', fontsize=11, fontweight='bold')
        ax.set_ylabel('Prognozuotos reikšmės', fontsize=11, fontweight='bold')
        ax.set_title(f'{indicator} (R² = {r2:.4f})', fontsize=12, fontweight='bold', pad=10)
        ax.legend(loc='upper left', fontsize=9)
        ax.grid(True, alpha=0.3, linestyle='--')

        # Add metrics text
        ax.text(0.025, 0.88, f'MAE: {mae:.2f}\nRMSE: {rmse:.2f}',
                transform=ax.transAxes, fontsize=9, verticalalignment='top',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))

    # Remove empty subplots
    for idx in range(n_indicators, len(axes)):
        fig.delaxes(axes[idx])

    plt.suptitle(f'{model_name} - Faktinių ir prognozuotų reikšmių palyginimas',
                 fontsize=16, fontweight='bold', y=0.995)
    plt.tight_layout()

    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=80, bbox_inches='tight')
    img_buffer.seek(0)
    plots['scatter_predictions'] = base64.b64encode(img_buffer.getvalue()).decode()
    plt.close()

    return plots

def generate_model_performance_plots(model_metrics):
    """Generate comprehensive model performance visualization plots"""
    plots = {}

    # Filter models with metrics (accept all model_type values that have nrmse or r2)
    # Possible model_type values: 'synthetic_test_no_zeros', 'no_test_nonzero', 'insufficient_data_mean_impute'
    regression_metrics = {col: metrics for col, metrics in model_metrics.items()
                         if metrics and ('nrmse' in metrics or 'r2' in metrics)}

    if not regression_metrics:
        print("[Plots] No regression metrics found to plot")
        return plots

    print(f"[Plots] Generating performance plots for {len(regression_metrics)} columns")

    # 1. Model Performance Summary Table Plot
    fig, ax = plt.subplots(figsize=(14, max(6, len(regression_metrics) * 0.8)))

    # Prepare data for table with normalized metrics
    columns = ['Stulpelis', 'nRMSE', 'R²', 'SMAPE (%)', 'nMAE', 'Imties dydis', 'Imtis (%)']
    table_data = []

    for col, metrics in regression_metrics.items():
        # Calculate sample percentage
        sample_percent = f"{(metrics['sample_size'] / metrics['total_samples'] * 100):.1f}" \
            if metrics.get('total_samples') and metrics.get('sample_size') else 'N/A'

        # Use normalized metrics
        nrmse_val = metrics.get('nrmse', float('nan'))
        r2_val = metrics.get('r2', float('nan'))
        smape_val = metrics.get('smape', float('nan'))
        nmae_val = metrics.get('nmae', float('nan'))

        table_data.append([
            col,
            f"{nrmse_val:.4f}" if not np.isnan(nrmse_val) else 'N/A',
            f"{r2_val:.4f}" if not np.isnan(r2_val) else 'N/A',
            f"{smape_val:.2f}" if not np.isnan(smape_val) else 'N/A',
            f"{nmae_val:.4f}" if not np.isnan(nmae_val) else 'N/A',
            str(metrics.get('sample_size', 'N/A')),
            sample_percent
        ])
    
    # Create table
    ax.axis('tight')
    ax.axis('off')
    
    table = ax.table(cellText=table_data,
                    colLabels=columns,
                    cellLoc='center',
                    loc='center',
                    bbox=[0, 0, 1, 1])
    
    # Style the table
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1, 2)
    
    # Header styling
    for i in range(len(columns)):
        table[(0, i)].set_facecolor('#4472C4')
        table[(0, i)].set_text_props(weight='bold', color='white')
    
    # Alternate row colors
    for i in range(1, len(table_data) + 1):
        for j in range(len(columns)):
            if i % 2 == 0:
                table[(i, j)].set_facecolor('#F2F2F2')
    
    plt.title('Modelių efektyvumo vertinimo metrikos', fontsize=14, fontweight='bold', pad=20)

    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=80, bbox_inches='tight')
    img_buffer.seek(0)
    plots['performance_table'] = base64.b64encode(img_buffer.getvalue()).decode()
    plt.close()
    
    # 2. nRMSE Comparison Bar Chart
    if len(regression_metrics) > 1:
        fig, ax = plt.subplots(figsize=(12, 8))

        columns = list(regression_metrics.keys())
        nrmse_values = [regression_metrics[col].get('nrmse', 0) for col in columns]

        # Filter out NaN values
        valid_data = [(col, val) for col, val in zip(columns, nrmse_values) if not np.isnan(val)]
        if valid_data:
            columns, nrmse_values = zip(*valid_data)
            columns = list(columns)
            nrmse_values = list(nrmse_values)

            bars = ax.bar(columns, nrmse_values, color=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd'][:len(columns)])

            ax.set_xlabel('Užpildyti stulpeliai', fontsize=12)
            ax.set_ylabel('nRMSE vertė', fontsize=12)
            ax.set_title('Normalized Root Mean Square Error (nRMSE) palyginimas', fontsize=14, fontweight='bold')

            # Add value labels on bars
            max_val = max(nrmse_values) if nrmse_values else 1
            for i, (bar, value) in enumerate(zip(bars, nrmse_values)):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max_val*0.01,
                       f'{value:.4f}', ha='center', va='bottom', fontsize=10, fontweight='bold')

            plt.xticks(rotation=45, ha='right')
            plt.grid(axis='y', alpha=0.3)
            plt.tight_layout()

            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=80, bbox_inches='tight')
            img_buffer.seek(0)
            plots['nrmse_comparison'] = base64.b64encode(img_buffer.getvalue()).decode()
            plt.close()
    
    # 3. R² Score Comparison
    if len(regression_metrics) > 1:
        fig, ax = plt.subplots(figsize=(12, 8))
        
        r2_values = [regression_metrics[col]['r2'] for col in columns]
        
        bars = ax.bar(columns, r2_values, color=['#2E8B57', '#FF6347', '#4682B4', '#DAA520', '#8A2BE2'][:len(columns)])
        
        ax.set_xlabel('Užpildyti stulpeliai', fontsize=12)
        ax.set_ylabel('R² koeficientas', fontsize=12)
        ax.set_title('Determinacijos koeficiento (R²) palyginimas', fontsize=14, fontweight='bold')
        ax.set_ylim(0, 1.1)
        
        # Add value labels on bars
        for i, (bar, value) in enumerate(zip(bars, r2_values)):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.02,
                   f'{value:.4f}', ha='center', va='bottom', fontsize=10, fontweight='bold')
        
        # Add interpretation lines
        ax.axhline(y=0.7, color='green', linestyle='--', alpha=0.7, label='Geras modelis (R² ≥ 0.7)')
        ax.axhline(y=0.5, color='orange', linestyle='--', alpha=0.7, label='Vidutinis modelis (R² ≥ 0.5)')
        ax.legend(loc='lower right')
        
        plt.xticks(rotation=45, ha='right')
        plt.grid(axis='y', alpha=0.3)
        plt.tight_layout()

        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=80, bbox_inches='tight')
        img_buffer.seek(0)
        plots['r2_comparison'] = base64.b64encode(img_buffer.getvalue()).decode()
        plt.close()
    
    # 4. SMAPE Comparison
    if len(regression_metrics) > 1:
        fig, ax = plt.subplots(figsize=(12, 8))

        columns = list(regression_metrics.keys())
        smape_values = [regression_metrics[col].get('smape', 0) for col in columns]

        # Filter out NaN values
        valid_data = [(col, val) for col, val in zip(columns, smape_values) if not np.isnan(val)]
        if valid_data:
            columns, smape_values = zip(*valid_data)
            columns = list(columns)
            smape_values = list(smape_values)

            bars = ax.bar(columns, smape_values, color=['#DC143C', '#FF8C00', '#32CD32', '#4169E1', '#8B008B'][:len(columns)])

            ax.set_xlabel('Užpildyti stulpeliai', fontsize=12)
            ax.set_ylabel('SMAPE (%)', fontsize=12)
            ax.set_title('Symmetric Mean Absolute Percentage Error (SMAPE) palyginimas', fontsize=14, fontweight='bold')

            # Add value labels on bars
            max_val = max(smape_values) if smape_values else 1
            for i, (bar, value) in enumerate(zip(bars, smape_values)):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max_val*0.02,
                       f'{value:.2f}%', ha='center', va='bottom', fontsize=10, fontweight='bold')

            # Add interpretation lines (SMAPE ranges 0-100%)
            ax.axhline(y=10, color='green', linestyle='--', alpha=0.7, label='Labai geras modelis (SMAPE ≤ 10%)')
            ax.axhline(y=20, color='orange', linestyle='--', alpha=0.7, label='Geras modelis (SMAPE ≤ 20%)')
            ax.legend(loc='upper right')

            plt.xticks(rotation=45, ha='right')
            plt.grid(axis='y', alpha=0.3)
            plt.tight_layout()
        
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=80, bbox_inches='tight')
            img_buffer.seek(0)
            plots['smape_comparison'] = base64.b64encode(img_buffer.getvalue()).decode()
            plt.close()
    
    # 5. Combined Metrics Radar Chart (if multiple models)
    if len(regression_metrics) > 1:
        fig, ax = plt.subplots(figsize=(10, 10), subplot_kw=dict(projection='polar'))

        # Normalize metrics for radar chart (0-1 scale)
        metrics_names = ['nRMSE', 'R²', 'SMAPE', 'nMAE']

        # Get valid metrics (filter out NaN values)
        valid_metrics = {col: m for col, m in regression_metrics.items()
                        if not np.isnan(m.get('nrmse', float('nan'))) and
                           not np.isnan(m.get('r2', float('nan')))}

        if valid_metrics:
            for i, (col, metrics) in enumerate(valid_metrics.items()):
                # Normalize values (invert nRMSE, SMAPE, nMAE so higher is better)
                max_nrmse = max([m.get('nrmse', 0) for m in valid_metrics.values()])
                max_smape = max([m.get('smape', 0) for m in valid_metrics.values()])
                max_nmae = max([m.get('nmae', 0) for m in valid_metrics.values()])

                normalized_values = [
                    1 - (metrics.get('nrmse', 0) / max_nrmse) if max_nrmse > 0 else 1,  # Inverted nRMSE
                    metrics.get('r2', 0),  # R² (already 0-1)
                    1 - (metrics.get('smape', 0) / max_smape) if max_smape > 0 else 1,  # Inverted SMAPE
                    1 - (metrics.get('nmae', 0) / max_nmae) if max_nmae > 0 else 1  # Inverted nMAE
                ]

                angles = np.linspace(0, 2 * np.pi, len(metrics_names), endpoint=False).tolist()
                normalized_values += normalized_values[:1]  # Complete the circle
                angles += angles[:1]

                color = plt.cm.Set1(i / len(valid_metrics))
                ax.plot(angles, normalized_values, 'o-', linewidth=2, label=col, color=color)
                ax.fill(angles, normalized_values, alpha=0.25, color=color)

            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(metrics_names)
            ax.set_ylim(0, 1)
            ax.set_title('Modelių efektyvumo palyginimas\n(Didesnis plotas = geresnis modelis)', y=1.08, fontsize=14, fontweight='bold')
            ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.0))

            plt.tight_layout()

            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=80, bbox_inches='tight')
            img_buffer.seek(0)
            plots['radar_comparison'] = base64.b64encode(img_buffer.getvalue()).decode()
            plt.close()
    
    return plots

@app.route('/api/system-status')
def get_system_status():
    """Get current system CPU and RAM usage"""
    try:
        import platform

        # CPU usage
        cpu_percent = psutil.cpu_percent(interval=0.1)
        cpu_count_logical = psutil.cpu_count(logical=True)  # Logical cores (with hyper-threading)
        cpu_count_physical = psutil.cpu_count(logical=False)  # Physical cores

        # Get CPU model/brand
        try:
            if platform.system() == 'Windows':
                import subprocess
                cpu_info = subprocess.check_output("wmic cpu get name", shell=True).decode()
                cpu_model = cpu_info.split('\n')[1].strip()
            else:
                # For Linux/Mac
                with open('/proc/cpuinfo', 'r') as f:
                    for line in f:
                        if 'model name' in line:
                            cpu_model = line.split(':')[1].strip()
                            break
                    else:
                        cpu_model = platform.processor()
        except:
            cpu_model = platform.processor() or "Nežinomas"

        # Memory usage
        memory = psutil.virtual_memory()
        memory_total = memory.total / (1024 ** 3)  # Convert to GB
        memory_used = memory.used / (1024 ** 3)
        memory_percent = memory.percent

        # Get RAM info - detailed modules information
        try:
            if platform.system() == 'Windows':
                import subprocess
                import re
                # Get detailed info using CSV format for better parsing
                ram_info = subprocess.check_output(
                    "wmic memorychip get manufacturer, partnumber, capacity, speed /format:csv",
                    shell=True
                ).decode()

                lines = [line.strip() for line in ram_info.split('\n') if line.strip()]
                # Remove header and empty lines
                lines = [line for line in lines if line and 'Node' not in line and 'Manufacturer' not in line]

                if lines:
                    modules = []
                    for line in lines:
                        try:
                            # CSV format: Node,Capacity,Manufacturer,PartNumber,Speed
                            parts = line.split(',')
                            if len(parts) >= 5:
                                capacity_str = parts[1].strip()
                                manufacturer = parts[2].strip()
                                partnumber = parts[3].strip()
                                speed = parts[4].strip()

                                # Parse capacity
                                if capacity_str.isdigit():
                                    capacity_bytes = int(capacity_str)
                                    capacity_gb = capacity_bytes / (1024 ** 3)

                                    # Build module info
                                    module_parts = []
                                    if manufacturer and manufacturer.lower() != 'unknown':
                                        module_parts.append(manufacturer)
                                    if partnumber:
                                        module_parts.append(partnumber)

                                    module_str = ' '.join(module_parts) if module_parts else "RAM"
                                    module_info = f"{module_str} {capacity_gb:.0f}GB"

                                    if speed and speed.isdigit():
                                        module_info += f" {speed}MHz"

                                    modules.append(module_info)
                        except (ValueError, IndexError) as e:
                            print(f"Error parsing RAM line: {line}, error: {e}")
                            continue

                    if modules:
                        # Show number of each module type
                        from collections import Counter
                        module_counts = Counter(modules)
                        module_strs = []
                        for module, count in module_counts.items():
                            if count > 1:
                                module_strs.append(f"{count}x {module}")
                            else:
                                module_strs.append(module)

                        if len(module_strs) == 1:
                            memory_model = module_strs[0]
                        else:
                            # Use + to show modules are summed together
                            memory_model = " + ".join(module_strs)
                    else:
                        memory_model = f"{memory_total:.0f} GB RAM"
                else:
                    memory_model = f"{memory_total:.0f} GB RAM"
            else:
                # For Linux (requires root for dmidecode)
                memory_model = f"{memory_total:.0f} GB RAM"
        except Exception as e:
            print(f"Error getting RAM info: {e}")
            import traceback
            traceback.print_exc()
            memory_model = f"{memory_total:.0f} GB RAM"

        # Disk usage (main drive)
        disk = psutil.disk_usage('/')
        disk_total = disk.total / (1024 ** 3)  # Convert to GB
        disk_used = disk.used / (1024 ** 3)
        disk_percent = disk.percent

        # Get disk info
        try:
            if platform.system() == 'Windows':
                import subprocess
                disk_info = subprocess.check_output("wmic diskdrive get model, size", shell=True).decode()
                lines = [line.strip() for line in disk_info.split('\n') if line.strip() and 'Model' not in line]
                if lines:
                    # Parse first disk
                    parts = lines[0].rsplit(None, 1)  # Split from right, last element is size
                    if len(parts) >= 2:
                        disk_model_name = parts[0].strip()
                        disk_size_bytes = int(parts[1])
                        disk_size_gb = disk_size_bytes / (1024 ** 3)
                        disk_model = f"{disk_model_name} ({disk_size_gb:.0f} GB)"
                    else:
                        disk_model = parts[0].strip() if parts else "Kietasis diskas"
                else:
                    disk_model = "Kietasis diskas"
            else:
                # For Linux
                partitions = psutil.disk_partitions()
                if partitions:
                    disk_model = partitions[0].device
                else:
                    disk_model = "Kietasis diskas"
        except:
            disk_model = "Kietasis diskas"

        return jsonify({
            'cpu': {
                'percent': round(cpu_percent, 1),
                'count_physical': cpu_count_physical,
                'count_logical': cpu_count_logical,
                'model': cpu_model
            },
            'memory': {
                'total': round(memory_total, 2),
                'used': round(memory_used, 2),
                'percent': round(memory_percent, 1),
                'model': memory_model
            },
            'disk': {
                'total': round(disk_total, 2),
                'used': round(disk_used, 2),
                'percent': round(disk_percent, 1),
                'model': disk_model
            }
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)